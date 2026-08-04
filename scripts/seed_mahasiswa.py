"""
Seed script — import data mahasiswa dari
"Pembagian_Kelompok_SPARK_SIC_Sept_2026.xlsx" (sheet "Sep26 Monitoring Daftar Ulang")
ke Supabase, sesuai skema baru (tanpa NIM, validasi nama + prodi).

Cara pakai:
  1. pip install openpyxl supabase --break-system-packages
  2. Set environment variables:
       export SUPABASE_URL="https://xxxx.supabase.co"
       export SUPABASE_SERVICE_ROLE_KEY="ey..."   <- WAJIB service role key, bukan anon key
     (service role key dibutuhkan karena RLS di tabel prodi/students hanya izinkan
     'authenticated'; seed script ini bukan proses login panitia, jadi harus bypass RLS)
  3. Sesuaikan EXCEL_PATH di bawah ke lokasi file di komputer kamu
  4. python3 seed_students.py
"""

import os
import re
import sys
from openpyxl import load_workbook
from supabase import create_client

# ------------------------------------------------------------------
# Konfigurasi
# ------------------------------------------------------------------
EXCEL_PATH = "data/Pembagian_Kelompok_SPARK_SIC_Sept_2026.xlsx"
SHEET_NAME = "Sep26 Monitoring Daftar Ulang"
DATA_START_ROW = 4  # baris 1-3 adalah header/grouping, data mulai baris 4

# Kolom di sheet (0-indexed, sesuai urutan values_only tuple)
COL_NAMA = 1
COL_PRODI = 3
COL_STATUS = 5
COL_EMAIL = 8
COL_NO_WA = 9

EXCLUDED_STATUS = {"cancel"}

# Mapping nama prodi (persis seperti di Excel) -> kode prodi.
# Kode ini yang dipakai sebagai UNIQUE key di tabel prodi.
PRODI_MAP = {
    "Accounting": "ACC",
    "AI": "AI",
    "Computer Science": "CS",
    "Data Science": "DS",
    "Desain Komunikasi Visual": "DKV",
    "Digital Business": "DB",
    "Finance": "FIN",
    "Ilmu Komunikasi": "IKOM",
    "Information System": "IS",
    "Law": "LAW",
    "Management": "MGT",
    "Pendidikan Guru Sekolah Dasar": "PGSD",
    "Psychology": "PSY",
    "Teknik Elektro": "TE",
    "Teknik Industri": "TI",
    "Teknik Lingkungan & Rekayasa Berkelanjutan": "TLRB",
}


def normalize_name(nama: str) -> str:
    """Samain persis dengan generated column nama_normalized di Postgres:
    lower(trim(regexp_replace(nama, '\\s+', ' ', 'g')))"""
    return re.sub(r"\s+", " ", nama.strip()).lower()


def normalize_wa(raw) -> str | None:
    if raw is None:
        return None
    digits = "".join(c for c in str(raw) if c.isdigit())
    return digits or None


def get_supabase_client():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit(
            "ERROR: SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY belum di-set. "
            "Lihat instruksi di bagian atas file ini."
        )
    return create_client(url, key)


def seed_prodi(supabase):
    """Upsert 16 prodi, return dict {nama_prodi_di_excel: prodi_id}"""
    rows = [{"kode": kode, "nama": nama} for nama, kode in PRODI_MAP.items()]
    result = supabase.table("prodi").upsert(rows, on_conflict="kode").execute()

    # Ambil ulang biar dapet id (upsert response kadang gak selalu return semua kolom)
    existing = supabase.table("prodi").select("id, kode").execute()
    kode_to_id = {row["kode"]: row["id"] for row in existing.data}

    nama_to_id = {}
    for nama_excel, kode in PRODI_MAP.items():
        if kode not in kode_to_id:
            print(f"  WARNING: prodi kode '{kode}' ({nama_excel}) gagal ke-insert")
            continue
        nama_to_id[nama_excel] = kode_to_id[kode]

    print(f"Prodi ready: {len(nama_to_id)}/16")
    return nama_to_id


def load_students_from_excel(prodi_lookup):
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    students = []
    skipped_cancel = 0
    skipped_blank = 0
    skipped_unknown_prodi = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, values_only=True):
        nama_raw = row[COL_NAMA]
        if not nama_raw or not str(nama_raw).strip():
            skipped_blank += 1
            continue

        status = str(row[COL_STATUS]).strip().lower() if row[COL_STATUS] else ""
        if status in EXCLUDED_STATUS:
            skipped_cancel += 1
            continue

        prodi_nama = row[COL_PRODI]
        prodi_id = prodi_lookup.get(prodi_nama)
        if not prodi_id:
            skipped_unknown_prodi += 1
            print(f"  WARNING: prodi '{prodi_nama}' tidak dikenal, skip: {nama_raw}")
            continue

        students.append(
            {
                "nama": str(nama_raw).strip(),
                "prodi_id": prodi_id,
                "email": (str(row[COL_EMAIL]).strip() if row[COL_EMAIL] else None),
                "no_wa": normalize_wa(row[COL_NO_WA]),
            }
        )

    print(
        f"Excel parsed: {len(students)} valid, "
        f"{skipped_cancel} status Cancel di-skip, "
        f"{skipped_blank} baris nama kosong di-skip, "
        f"{skipped_unknown_prodi} prodi tidak dikenal di-skip"
    )
    return students


def insert_students(supabase, students, batch_size=500):
    inserted = 0
    duplicates = 0

    for i in range(0, len(students), batch_size):
        batch = students[i : i + batch_size]
        result = (
            supabase.table("mahasiswa")
            .upsert(batch, on_conflict="nama_normalized,prodi_id", ignore_duplicates=True)
            .execute()
        )
        inserted += len(result.data)
        duplicates += len(batch) - len(result.data)
        print(f"  Batch {i // batch_size + 1}: {len(result.data)}/{len(batch)} ter-insert")

    print(f"\nTotal ter-insert: {inserted}")
    print(f"Total duplikat (nama+prodi sama persis, di-skip): {duplicates}")


def main():
    print("=== Seed Students — PKKMB System ===\n")
    supabase = get_supabase_client()

    print("1. Seeding prodi...")
    prodi_lookup = seed_prodi(supabase)

    print("\n2. Membaca data mahasiswa dari Excel...")
    students = load_students_from_excel(prodi_lookup)

    print("\n3. Insert ke tabel students...")
    insert_students(supabase, students)

    print("\nSelesai.")


if __name__ == "__main__":
    main()